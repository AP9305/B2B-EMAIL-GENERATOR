File path: mytabs.py

New file content:
import pandas as pd
import openpyxl

# Sample data
data = {
    'company_name': [
        'TechCorp Solutions',
        'Innovate Labs',
        'Global Systems Inc',
        'Future Tech',
        'Smart Solutions'
    ],
    'head_name': [
        'Michael Chen',
        'Sarah Patel',
        'David Wilson',
        'Emily Brown',
        'Raj Kumar'
    ],
    'head_email': [
        'michael.chen@techcorp.com',
        'sarah.patel@innovatelabs.com',
        'david.wilson@globalsystems.com',
        'emily.brown@futuretech.com',
        'raj.kumar@smartsolutions.com'
    ],
    'website': [
        'https://techcorp.com',
        'https://innovatelabs.com',
        'https://globalsystems.com',
        'https://futuretech.com',
        'https://smartsolutions.com'
    ],
    'linkedin_url': [
        'https://linkedin.com/in/michaelchen',
        'https://linkedin.com/in/sarahpatel',
        'https://linkedin.com/in/davidwilson',
        'https://linkedin.com/in/emilybrown',
        'https://linkedin.com/in/rajkumar'
    ]
}

# Create DataFrame
df = pd.DataFrame(data)

# Save to Excel with explicit engine
try:
    df.to_excel('sample.xlsx', index=False, engine='openpyxl')
    print("Sample Excel file created successfully!")
except Exception as e:
    print(f"Error creating Excel file: {str(e)}")
    
    # Try alternative method
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Write headers
        for col, header in enumerate(df.columns, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        wb.save('sample.xlsx')
        print("Sample Excel file created successfully using alternative method!")
    except Exception as e:
        print(f"Error creating Excel file using alternative method: {str(e)}")